"""Генератор файла финансовой модели (.xlsx).

Универсальная мультииндустриальная инвестиционная модель: 36 месяцев,
живые формулы (ничего не захардкожено), сценарии Пессимист/База/Оптимист,
сезонность, кредит с аннуитетом, NPV/IRR/окупаемость, матрица
чувствительности и Excel-диаграммы. Все ключевые входы имеют именованные
диапазоны — это упрощает и ручную работу, и работу ИИ-агента в Excel.

Цветовой стандарт финмоделей: синий шрифт — входные данные,
чёрный — формулы, зелёный — ссылки между листами.
"""
import io
from datetime import date

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

HORIZON = 36
FIRST_COL = 3  # колонка C — первый месяц

F = "Arial"
INPUT = Font(name=F, color="0000FF", size=10)              # входные данные
INPUT_B = Font(name=F, color="0000FF", size=10, bold=True)
FORMULA = Font(name=F, size=10)                            # формулы
LINK = Font(name=F, color="008000", size=10)               # ссылки между листами
GREY = Font(name=F, color="8A8F8C", size=9)
H1 = Font(name=F, size=16, bold=True)
H2 = Font(name=F, size=11, bold=True)
BOLD = Font(name=F, size=10, bold=True)
WHITE_B = Font(name=F, size=10, bold=True, color="FFFFFF")

FILL_INPUT = PatternFill("solid", start_color="FFF2CC")
FILL_HEAD = PatternFill("solid", start_color="14604D")
FILL_SECTION = PatternFill("solid", start_color="E9F1ED")
FILL_TOTAL = PatternFill("solid", start_color="F2F4F2")
THIN = Side(style="thin", color="D9DDD9")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

MONEY = '# ##0\u00a0₽;(# ##0\u00a0₽);"–"'
NUM = '# ##0;(# ##0);"–"'
PCT = "0.0%"
COEF = "0.00"


def _name(wb: Workbook, name: str, sheet: str, ref: str) -> None:
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def _input(ws, cell: str, value, fmt: str | None = None, bold: bool = False):
    c = ws[cell]
    c.value = value
    c.font = INPUT_B if bold else INPUT
    c.fill = FILL_INPUT
    c.border = BORDER
    if fmt:
        c.number_format = fmt
    return c


def _label(ws, cell: str, text: str, font: Font = FORMULA):
    ws[cell] = text
    ws[cell].font = font


def _section(ws, row: int, text: str, width: int = 6):
    ws.cell(row=row, column=1, value=text).font = H2
    for col in range(1, width + 1):
        ws.cell(row=row, column=col).fill = FILL_SECTION


def build_model_workbook(title: str = "Новый проект", drivers: dict | None = None) -> Workbook:
    """Собирает книгу модели. drivers — стартовые значения входов (опционально)."""
    d = drivers or {}
    wb = Workbook()

    # ============================================================ ОБЛОЖКА ====
    ws = wb.active
    ws.title = "Обложка"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 70
    ws["B3"] = title
    ws["B3"].font = Font(name=F, size=24, bold=True, color="14604D")
    _label(ws, "B4", "Финансовая модель · инвестиционный формат", Font(name=F, size=12, color="66726B"))
    _label(ws, "B6", f"Дата создания: {date.today().strftime('%d.%m.%Y')} · Версия 1.0 · Горизонт: {HORIZON} месяцев", GREY)
    _label(ws, "B8", "Как читать модель", H2)
    legend = [
        ("СИНИЙ шрифт на жёлтом фоне — входные данные: меняйте их", INPUT),
        ("ЧЁРНЫЙ шрифт — формулы: не редактируйте без необходимости", FORMULA),
        ("ЗЕЛЁНЫЙ шрифт — ссылки на другие листы книги", LINK),
    ]
    for i, (text, font) in enumerate(legend):
        _label(ws, f"B{9 + i}", text, font)
    _label(ws, "B13", "Состав книги", H2)
    toc = [
        "Допущения — все входы модели: сценарии, выручка, расходы, CAPEX, кредит",
        "Модель — помесячный расчёт на 36 месяцев: P&L, кредит, денежные потоки",
        "Годовая сводка — агрегаты по годам модели",
        "Дашборд — KPI (NPV, IRR, окупаемость) и диаграммы",
        "Чувствительность — матрица NPV: выручка × себестоимость",
        "Инструкция — как работать с моделью и с ИИ-агентом в Excel",
    ]
    for i, line in enumerate(toc):
        _label(ws, f"B{14 + i}", "· " + line, FORMULA)

    # =========================================================== ДОПУЩЕНИЯ ===
    A = wb.create_sheet("Допущения")
    A.sheet_view.showGridLines = False
    A.column_dimensions["A"].width = 38
    for col in "BCDE":
        A.column_dimensions[col].width = 14
    A.column_dimensions["F"].width = 52

    _label(A, "A1", "Допущения модели", H1)
    _label(A, "F1", "Комментарий", GREY)

    _section(A, 3, "Общие")
    _label(A, "A4", "Название проекта")
    _input(A, "B4", title)
    _label(A, "A5", "Старт модели (первое число месяца)")
    _input(A, "B5", d.get("start_date") or date.today().replace(day=1), "DD.MM.YYYY")
    _label(A, "A6", "Ставка дисконтирования, % годовых")
    _input(A, "B6", d.get("discount_rate", 0.20), PCT)
    _label(A, "A7", "Месячная ставка (расчёт)")
    A["B7"] = "=(1+B6)^(1/12)-1"
    A["B7"].font = FORMULA
    A["B7"].number_format = "0.000%"
    _label(A, "A8", "Эффективная ставка налога, % от выручки")
    _input(A, "B8", d.get("tax_rate", 0.06), PCT)
    _label(A, "F8", "Упрощение: налог считается от выручки. Для УСН-15/ОСНО скорректируйте формулу налога на листе «Модель».", GREY)

    _section(A, 10, "Сценарий (анализ чувствительности «на лету»)")
    _label(A, "A11", "Активный сценарий")
    sc = _input(A, "B11", "База", bold=True)
    sc.alignment = Alignment(horizontal="center")
    dv = DataValidation(type="list", formula1='"Пессимист,База,Оптимист"', allow_blank=False)
    A.add_data_validation(dv)
    dv.add(A["B11"])
    _label(A, "F11", "Выпадающий список: меняете сценарий — пересчитывается вся модель.", GREY)
    _label(A, "A12", "Отклонение выручки по сценариям")
    _label(A, "B13", "Пессимист", GREY); _label(A, "C13", "База", GREY); _label(A, "D13", "Оптимист", GREY)
    _input(A, "B14", d.get("rev_pess", -0.15), PCT)
    _input(A, "C14", 0.0, PCT)
    _input(A, "D14", d.get("rev_opt", 0.15), PCT)
    _label(A, "A15", "Отклонение себестоимости по сценариям")
    _input(A, "B16", d.get("cost_pess", 0.10), PCT)
    _input(A, "C16", 0.0, PCT)
    _input(A, "D16", d.get("cost_opt", -0.05), PCT)
    _label(A, "A17", "Активное отклонение выручки")
    A["B17"] = '=IF(B11="Пессимист",B14,IF(B11="Оптимист",D14,C14))'
    A["B17"].font = FORMULA; A["B17"].number_format = PCT
    _label(A, "A18", "Активное отклонение себестоимости")
    A["B18"] = '=IF(B11="Пессимист",B16,IF(B11="Оптимист",D16,C16))'
    A["B18"].font = FORMULA; A["B18"].number_format = PCT

    _section(A, 20, "Выручка")
    _label(A, "A21", "Базовая выручка в месяц, ₽")
    _input(A, "B21", d.get("base_revenue", 1_000_000), MONEY)
    _label(A, "A22", "Рост выручки, % месяц к месяцу")
    _input(A, "B22", d.get("growth_mom", 0.02), PCT)
    _label(A, "A23", "Разгон до полной мощности, мес. (0 — сразу)")
    _input(A, "B23", d.get("ramp_months", 0), NUM)
    _label(A, "A25", "Сезонность по месяцам года (1.00 = без сезонности)")
    months_ru = ["Янв", "Фев", "Мар", "Апр", "Май", "Июн", "Июл", "Авг", "Сен", "Окт", "Ноя", "Дек"]
    for i, name in enumerate(months_ru):
        col = get_column_letter(2 + i)
        _label(A, f"{col}26", name, GREY)
        A[f"{col}26"].alignment = Alignment(horizontal="center")
        _input(A, f"{col}27", (d.get("seasonality") or [1.0] * 12)[i], COEF).alignment = Alignment(horizontal="center")

    _section(A, 29, "Операционные расходы")
    _label(A, "A30", "Себестоимость (COGS), % от выручки")
    _input(A, "B30", d.get("cogs_pct", 0.35), PCT)
    _label(A, "A31", "ФОТ с налогами, ₽ в месяц")
    _input(A, "B31", d.get("payroll", 300_000), MONEY)
    _label(A, "A32", "Аренда и коммунальные, ₽ в месяц")
    _input(A, "B32", d.get("rent", 100_000), MONEY)
    _label(A, "A33", "Маркетинг, % от выручки")
    _input(A, "B33", d.get("marketing_pct", 0.07), PCT)
    _label(A, "A34", "Прочие расходы, % от выручки")
    _input(A, "B34", d.get("other_pct", 0.03), PCT)
    _label(A, "A35", "Индексация ФОТ и аренды, % в год")
    _input(A, "B35", d.get("index_rate", 0.08), PCT)

    _section(A, 37, "Инвестиции (CAPEX)")
    _label(A, "A38", "Сумма стартовых вложений, ₽")
    _input(A, "B38", d.get("capex_total", 0), MONEY)
    _label(A, "A39", "Распределение CAPEX, месяцев")
    _input(A, "B39", d.get("capex_months", 6), NUM)

    _section(A, 41, "Финансирование (кредит)")
    _label(A, "A42", "Сумма кредита, ₽ (0 — без кредита)")
    _input(A, "B42", d.get("loan_amount", 0), MONEY)
    _label(A, "A43", "Ставка по кредиту, % годовых")
    _input(A, "B43", d.get("loan_rate", 0.17), PCT)
    _label(A, "A44", "Срок кредита, месяцев")
    _input(A, "B44", d.get("loan_term", 60), NUM)
    _label(A, "A45", "Месяц получения кредита (№)")
    _input(A, "B45", d.get("loan_month", 1), NUM)

    names = {
        "StartDate": "$B$5", "DiscountRate": "$B$6", "MonthlyRate": "$B$7", "TaxRate": "$B$8",
        "RevAdj": "$B$17", "CostAdj": "$B$18",
        "BaseRevenue": "$B$21", "GrowthMoM": "$B$22", "RampMonths": "$B$23",
        "Seasonality": "$B$27:$M$27",
        "CogsPct": "$B$30", "Payroll": "$B$31", "Rent": "$B$32",
        "MarketingPct": "$B$33", "OtherPct": "$B$34", "IndexRate": "$B$35",
        "CapexTotal": "$B$38", "CapexMonths": "$B$39",
        "LoanAmount": "$B$42", "LoanRate": "$B$43", "LoanTerm": "$B$44", "LoanMonth": "$B$45",
    }
    for nm, ref in names.items():
        _name(wb, nm, "Допущения", ref)

    # ============================================================== МОДЕЛЬ ===
    M = wb.create_sheet("Модель")
    M.column_dimensions["A"].width = 34
    M.column_dimensions["B"].width = 3
    last_col = FIRST_COL + HORIZON - 1
    for c in range(FIRST_COL, last_col + 1):
        M.column_dimensions[get_column_letter(c)].width = 11
    M.freeze_panes = "C5"

    _label(M, "A1", "Помесячная модель", H1)
    _label(M, "A2", "Расходы показаны положительными числами; итоги их вычитают.", GREY)

    for c in range(FIRST_COL, last_col + 1):
        col = get_column_letter(c)
        t = c - FIRST_COL + 1
        M[f"{col}3"] = t
        M[f"{col}3"].font = GREY
        M[f"{col}3"].alignment = Alignment(horizontal="center")
        M[f"{col}4"] = f"=EDATE(StartDate,{col}3-1)" if c > FIRST_COL else "=StartDate"
        M[f"{col}4"].font = BOLD
        M[f"{col}4"].number_format = "MMM YY"
        M[f"{col}4"].alignment = Alignment(horizontal="center")
        M[f"{col}4"].fill = FILL_TOTAL
    _label(M, "A3", "№ месяца", GREY)
    _label(M, "A4", "Месяц", BOLD)

    rows: dict[str, int] = {}

    def put_row(r: int, label: str, formula: str, fmt: str = MONEY, font: Font = FORMULA,
                bold: bool = False, fill=None, key: str | None = None):
        M.cell(row=r, column=1, value=label).font = BOLD if bold else FORMULA
        if fill:
            M.cell(row=r, column=1).fill = fill
        for c in range(FIRST_COL, last_col + 1):
            col = get_column_letter(c)
            cell = M[f"{col}{r}"]
            cell.value = formula.replace("{c}", col).replace("{p}", get_column_letter(c - 1))
            cell.number_format = fmt
            cell.font = Font(name=F, size=10, bold=True) if bold else font
            if fill:
                cell.fill = fill
        if key:
            rows[key] = r

    r = 6
    _section(M, r, "ВЫРУЧКА", 2); r += 1
    put_row(r, "Сезонный коэффициент", "=INDEX(Seasonality,MONTH({c}$4))", COEF, GREY, key="season"); r += 1
    put_row(r, "Выручка",
            "=BaseRevenue*(1+GrowthMoM)^({c}$3-1)*IF(RampMonths=0,1,MIN(1,{c}$3/RampMonths))"
            f"*{{c}}{rows['season']}*(1+RevAdj)",
            bold=True, fill=FILL_TOTAL, key="rev"); r += 2

    _section(M, r, "ОПЕРАЦИОННЫЕ РАСХОДЫ", 2); r += 1
    put_row(r, "Себестоимость (COGS)", f"={{c}}{rows['rev']}*CogsPct*(1+CostAdj)", key="cogs"); r += 1
    put_row(r, "Валовая прибыль", f"={{c}}{rows['rev']}-{{c}}{rows['cogs']}", bold=True, key="gross"); r += 1
    put_row(r, "ФОТ с налогами", "=Payroll*(1+IndexRate)^INT(({c}$3-1)/12)", key="payroll"); r += 1
    put_row(r, "Аренда и коммунальные", "=Rent*(1+IndexRate)^INT(({c}$3-1)/12)", key="rent"); r += 1
    put_row(r, "Маркетинг", f"={{c}}{rows['rev']}*MarketingPct", key="mkt"); r += 1
    put_row(r, "Прочие расходы", f"={{c}}{rows['rev']}*OtherPct", key="other"); r += 1
    put_row(r, "Итого операционные расходы",
            f"=SUM({{c}}{rows['payroll']}:{{c}}{rows['other']})", bold=True, key="opex"); r += 1
    put_row(r, "EBITDA", f"={{c}}{rows['gross']}-{{c}}{rows['opex']}", bold=True, fill=FILL_TOTAL, key="ebitda"); r += 1
    put_row(r, "Налог", f"=MAX(0,{{c}}{rows['rev']}*TaxRate)", key="tax"); r += 1
    put_row(r, "Чистая прибыль", f"={{c}}{rows['ebitda']}-{{c}}{rows['tax']}",
            bold=True, fill=FILL_TOTAL, key="net"); r += 2

    _section(M, r, "ИНВЕСТИЦИИ", 2); r += 1
    put_row(r, "CAPEX", "=IF({c}$3<=CapexMonths,CapexTotal/MAX(CapexMonths,1),0)", key="capex"); r += 2

    _section(M, r, "КРЕДИТ", 2); r += 1
    put_row(r, "Получение кредита", "=IF({c}$3=LoanMonth,LoanAmount,0)", key="loan_in"); r += 1
    put_row(r, "Платёж по кредиту (аннуитет)",
            "=IF(AND(LoanAmount>0,{c}$3>=LoanMonth,{c}$3<LoanMonth+LoanTerm),"
            "-PMT(LoanRate/12,LoanTerm,LoanAmount),0)", key="loan_pay"); r += 1
    put_row(r, "в т.ч. проценты",
            "=IF(AND(LoanAmount>0,{c}$3>=LoanMonth,{c}$3<LoanMonth+LoanTerm),"
            "-IPMT(LoanRate/12,{c}$3-LoanMonth+1,LoanTerm,LoanAmount),0)", font=GREY, key="loan_int"); r += 1
    put_row(r, "Остаток долга",
            f"=MAX(0,IF({{c}}$3<LoanMonth,0,IF({{c}}$3=LoanMonth,LoanAmount,{{p}}{r})"
            f"-({{c}}{rows['loan_pay']}-{{c}}{rows['loan_int']})))", font=GREY, key="loan_bal"); r += 2

    _section(M, r, "ДЕНЕЖНЫЙ ПОТОК", 2); r += 1
    put_row(r, "Операционный поток", f"={{c}}{rows['net']}", key="cf_op"); r += 1
    put_row(r, "Инвестиционный поток", f"=-{{c}}{rows['capex']}", key="cf_inv"); r += 1
    put_row(r, "Финансовый поток", f"={{c}}{rows['loan_in']}-{{c}}{rows['loan_pay']}", key="cf_fin"); r += 1
    put_row(r, "FCF проекта (без финансирования)", f"={{c}}{rows['cf_op']}+{{c}}{rows['cf_inv']}",
            bold=True, fill=FILL_TOTAL, key="fcf"); r += 1
    put_row(r, "FCF накопленным итогом",
            f"=IF({{c}}$3=1,{{c}}{rows['fcf']},{{p}}{r}+{{c}}{rows['fcf']})", key="fcf_cum"); r += 1
    put_row(r, "Чистый денежный поток", f"={{c}}{rows['cf_op']}+{{c}}{rows['cf_inv']}+{{c}}{rows['cf_fin']}",
            key="cf_net"); r += 1
    put_row(r, "Остаток денежных средств",
            f"=IF({{c}}$3=1,{{c}}{rows['cf_net']},{{p}}{r}+{{c}}{rows['cf_net']})",
            bold=True, fill=FILL_TOTAL, key="cash"); r += 1
    put_row(r, "Коэффициент дисконтирования", "=1/(1+MonthlyRate)^{c}$3", COEF, GREY, key="df")

    first_l, last_l = get_column_letter(FIRST_COL), get_column_letter(last_col)
    for key in ("rev", "cogs", "fcf", "fcf_cum", "df", "ebitda", "cf_net", "cash"):
        _name(wb, {"rev": "RevenueRow", "cogs": "CogsRow", "fcf": "FCF", "fcf_cum": "CumFCF",
                   "df": "DFRow", "ebitda": "EbitdaRow", "cf_net": "NetCFRow", "cash": "CashRow"}[key],
              "Модель", f"${first_l}${rows[key]}:${last_l}${rows[key]}")

    # ======================================================= ГОДОВАЯ СВОДКА ==
    Y = wb.create_sheet("Годовая сводка")
    Y.sheet_view.showGridLines = False
    Y.column_dimensions["A"].width = 34
    for col in "BCD":
        Y.column_dimensions[col].width = 18
    _label(Y, "A1", "Годовая сводка (по годам модели)", H1)
    year_ranges = []
    for y in range(3):
        c0 = get_column_letter(FIRST_COL + y * 12)
        c1 = get_column_letter(FIRST_COL + y * 12 + 11)
        year_ranges.append((c0, c1))
        col = get_column_letter(2 + y)
        Y[f"{col}3"] = f"Год {y + 1}"
        Y[f"{col}3"].font = WHITE_B
        Y[f"{col}3"].fill = FILL_HEAD
        Y[f"{col}3"].alignment = Alignment(horizontal="center")
    y_items = [("Выручка", "rev", MONEY), ("Себестоимость", "cogs", MONEY), ("Валовая прибыль", "gross", MONEY),
               ("Операционные расходы", "opex", MONEY), ("EBITDA", "ebitda", MONEY),
               ("Налог", "tax", MONEY), ("Чистая прибыль", "net", MONEY),
               ("CAPEX", "capex", MONEY), ("FCF проекта", "fcf", MONEY)]
    for i, (label, key, fmt) in enumerate(y_items):
        rr = 4 + i
        _label(Y, f"A{rr}", label, BOLD if key in ("gross", "ebitda", "net", "fcf") else FORMULA)
        for y in range(3):
            c0, c1 = year_ranges[y]
            cell = Y[f"{get_column_letter(2 + y)}{rr}"]
            cell.value = f"=SUM(Модель!{c0}{rows[key]}:{c1}{rows[key]})"
            cell.number_format = fmt
            cell.font = LINK
    rr = 4 + len(y_items)
    _label(Y, f"A{rr}", "EBITDA маржа", BOLD)
    for y in range(3):
        col = get_column_letter(2 + y)
        Y[f"{col}{rr}"] = f"=IF({col}4=0,0,{col}8/{col}4)"
        Y[f"{col}{rr}"].number_format = PCT
        Y[f"{col}{rr}"].font = FORMULA

    # ============================================================= ДАШБОРД ===
    D = wb.create_sheet("Дашборд")
    D.sheet_view.showGridLines = False
    D.column_dimensions["A"].width = 2
    for col in "BDFH":
        D.column_dimensions[col].width = 24
    for col in "CEGI":
        D.column_dimensions[col].width = 6
    _label(D, "B1", "Дашборд проекта", H1)
    D["B2"] = "='Допущения'!B4"
    D["B2"].font = Font(name=F, size=12, color="66726B")

    kpi = [
        ("B4", "NPV проекта", "=SUMPRODUCT(FCF,DFRow)", MONEY),
        ("D4", "IRR, % годовых", '=IFERROR((1+IRR(FCF))^12-1,"н/д")', PCT),
        ("F4", "Окупаемость, мес.",
         f'=IF(COUNTIF(CumFCF,"<0")=0,0,IF(COUNTIF(CumFCF,"<0")={HORIZON},"за горизонтом",COUNTIF(CumFCF,"<0")+1))', NUM),
        ("H4", "Остаток денег, конец горизонта", f"=Модель!{last_l}{rows['cash']}", MONEY),
        ("B7", "Выручка, год 1", "='Годовая сводка'!B4", MONEY),
        ("D7", "EBITDA, год 1", "='Годовая сводка'!B8", MONEY),
        ("F7", "EBITDA маржа, год 1", "='Годовая сводка'!B13", PCT),
        ("H7", "Чистая прибыль, год 1", "='Годовая сводка'!B10", MONEY),
    ]
    for cell, label, formula, fmt in kpi:
        D[cell] = label
        D[cell].font = GREY
        v = D[f"{cell[0]}{int(cell[1:]) + 1}"]
        v.value = formula
        v.number_format = fmt
        v.font = Font(name=F, size=14, bold=True, color="14604D")

    dates_ref = Reference(M, min_col=FIRST_COL, max_col=last_col, min_row=4, max_row=4)

    ch1 = LineChart()
    ch1.title = "Выручка и EBITDA по месяцам"
    ch1.height, ch1.width = 8, 22
    ch1.add_data(Reference(M, min_col=FIRST_COL - 1, max_col=last_col, min_row=rows["rev"], max_row=rows["rev"]), titles_from_data=True)
    ch1.add_data(Reference(M, min_col=FIRST_COL - 1, max_col=last_col, min_row=rows["ebitda"], max_row=rows["ebitda"]), titles_from_data=True)
    ch1.set_categories(dates_ref)
    D.add_chart(ch1, "B10")

    ch2 = BarChart()
    ch2.title = "Чистый денежный поток"
    ch2.height, ch2.width = 8, 10.5
    ch2.add_data(Reference(M, min_col=FIRST_COL - 1, max_col=last_col, min_row=rows["cf_net"], max_row=rows["cf_net"]), titles_from_data=True)
    ch2.set_categories(dates_ref)
    ch2.legend = None
    D.add_chart(ch2, "B27")

    ch3 = LineChart()
    ch3.title = "FCF накопленным итогом и остаток денег"
    ch3.height, ch3.width = 8, 10.5
    ch3.add_data(Reference(M, min_col=FIRST_COL - 1, max_col=last_col, min_row=rows["fcf_cum"], max_row=rows["fcf_cum"]), titles_from_data=True)
    ch3.add_data(Reference(M, min_col=FIRST_COL - 1, max_col=last_col, min_row=rows["cash"], max_row=rows["cash"]), titles_from_data=True)
    ch3.set_categories(dates_ref)
    D.add_chart(ch3, "F27")

    # ====================================================== ЧУВСТВИТЕЛЬНОСТЬ =
    S = wb.create_sheet("Чувствительность")
    S.sheet_view.showGridLines = False
    S.column_dimensions["A"].width = 26
    for c in range(2, 8):
        S.column_dimensions[get_column_letter(c)].width = 15
    _label(S, "A1", "Чувствительность NPV", H1)
    _label(S, "A2", "Строки — дополнительное отклонение выручки; столбцы — себестоимости. Применяется поверх активного сценария.", GREY)

    _label(S, "A4", "NPV базовый (текущий сценарий)")
    S["B4"] = "=SUMPRODUCT(FCF,DFRow)"
    S["B4"].font = LINK; S["B4"].number_format = MONEY
    _label(S, "A5", "Чувствительность к выручке, ₽ на +1 п.п.")
    S["B5"] = "=SUMPRODUCT(RevenueRow,DFRow)*(1-CogsPct*(1+CostAdj)-MarketingPct-OtherPct-TaxRate)/100"
    S["B5"].font = FORMULA; S["B5"].number_format = MONEY
    _label(S, "A6", "Чувствительность к COGS, ₽ на +1 п.п.")
    S["B6"] = "=-SUMPRODUCT(CogsRow,DFRow)/100"
    S["B6"].font = FORMULA; S["B6"].number_format = MONEY

    rev_steps = [-0.20, -0.10, 0.0, 0.10, 0.20]
    cost_steps = [-0.10, -0.05, 0.0, 0.05, 0.10]
    base_r = 9
    _label(S, f"A{base_r - 1}", "Матрица NPV", H2)
    S[f"A{base_r}"] = "Δ выручки \\ Δ COGS"
    S[f"A{base_r}"].font = GREY
    for j, cstep in enumerate(cost_steps):
        cell = S.cell(row=base_r, column=2 + j, value=cstep)
        cell.font = INPUT; cell.fill = FILL_INPUT; cell.number_format = PCT
        cell.alignment = Alignment(horizontal="center")
    for i, rstep in enumerate(rev_steps):
        cell = S.cell(row=base_r + 1 + i, column=1, value=rstep)
        cell.font = INPUT; cell.fill = FILL_INPUT; cell.number_format = PCT
        for j in range(len(cost_steps)):
            col = get_column_letter(2 + j)
            target = S.cell(row=base_r + 1 + i, column=2 + j)
            target.value = f"=$B$4+$A{base_r + 1 + i}*100*$B$5+{col}${base_r}*100*$B$6"
            target.number_format = MONEY
            target.font = FORMULA
            target.border = BORDER
    matrix = f"B{base_r + 1}:{get_column_letter(1 + len(cost_steps))}{base_r + len(rev_steps)}"
    S.conditional_formatting.add(matrix, ColorScaleRule(
        start_type="min", start_color="E8B4A8",
        mid_type="num", mid_value=0, mid_color="FFFFFF",
        end_type="max", end_color="9CC3B4"))
    _label(S, f"A{base_r + len(rev_steps) + 2}",
           "Матрица линейна по выручке и COGS — это точное свойство модели (налог, маркетинг и прочие заданы долями от выручки).", GREY)

    # =========================================================== ИНСТРУКЦИЯ ==
    I = wb.create_sheet("Инструкция")
    I.sheet_view.showGridLines = False
    I.column_dimensions["A"].width = 4
    I.column_dimensions["B"].width = 100
    guide = [
        ("Как работать с моделью", H2),
        ("1. Все входы — на листе «Допущения» (синие ячейки на жёлтом фоне). Меняйте только их.", FORMULA),
        ("2. Сценарии: ячейка «Активный сценарий» — выпадающий список. Границы интервалов (Пессимист/Оптимист) тоже редактируются.", FORMULA),
        ("3. Сезонность: 12 коэффициентов. Пример для общепита: лето 1.2, январь 0.7.", FORMULA),
        ("4. Лист «Модель» — формулы. Добавляя строки, копируйте формулу на все 36 месяцев.", FORMULA),
        ("5. «Чувствительность» — матрица NPV с подсветкой; шаги отклонений редактируемы (синие).", FORMULA),
        ("", FORMULA),
        ("Мультииндустриальные пресеты (ориентиры для «Допущений»)", H2),
        ("Розница / e-com: COGS 55%, маркетинг 8%, ФОТ фикс, сезонность по пикам продаж.", FORMULA),
        ("Общепит: COGS (фудкост) 35%, ФОТ высокий, сезонность лето/праздники.", FORMULA),
        ("Услуги / агентство: COGS 0–10%, ФОТ — главная статья (40–50% выручки → задайте фикс).", FORMULA),
        ("SaaS / подписки: COGS 10–15%, маркетинг 15%+, рост м/м 3–8%, разгон 6–12 мес.", FORMULA),
        ("Производство / опт: COGS 50%, логистика в прочих, индексация 8–10%.", FORMULA),
        ("Девелопмент / агро: CAPEX и кредит — ключевые блоки; выручка с разгоном после ввода объекта.", FORMULA),
        ("", FORMULA),
        ("Работа с ИИ-агентом прямо в Excel", H2),
        ("Ключевые ячейки и ряды имеют ИМЕНА: BaseRevenue, CogsPct, FCF, CumFCF, RevenueRow, DFRow и др.", FORMULA),
        ("Просите агента ссылаться на имена: «увеличь GrowthMoM до 3% и скажи, как изменится NPV»,", FORMULA),
        ("«добавь строку расходов на лицензии 2% от RevenueRow ниже маркетинга», «поясни формулу аннуитета».", FORMULA),
        ("Ваши правки модели — обучающий материал: фиксируйте, что меняете, эти паттерны лягут в базу платформы.", FORMULA),
    ]
    for i, (text, font) in enumerate(guide):
        _label(I, f"B{2 + i}", text, font)

    wb.active = 0
    return wb


def export_model_bytes(title: str, drivers: dict | None = None) -> bytes:
    wb = build_model_workbook(title=title, drivers=drivers)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
