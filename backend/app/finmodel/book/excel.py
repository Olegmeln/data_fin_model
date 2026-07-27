"""Экспорт книги в Excel по реестру листов.

Конвенции финмоделей: синий шрифт — вводы, чёрный — формулы, зелёный —
ссылки на другой лист; итоги — живыми формулами (SUM/ссылки), а не
зашитыми числами. Помесячные ряды операционного движка пишутся
значениями с явной пометкой источника «движок AFM&C» (формульная
трассировка рядов — этап 5.3).
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

from ..assumptions_schema import AssumptionSet
from .engine import BookData, npv_sensitivity
from .registry import SheetSpec, resolve_sheets

FONT = "Arial"
INPUT = Font(name=FONT, color="0000FF")            # синий — ввод
FORMULA = Font(name=FONT, color="000000")          # чёрный — формула
LINK = Font(name=FONT, color="008000")             # зелёный — ссылка на лист
HEADER = Font(name=FONT, bold=True)
TITLE = Font(name=FONT, bold=True, size=14)
NOTE = Font(name=FONT, italic=True, size=9, color="666666")
KEY_FILL = PatternFill("solid", start_color="FFFF00")
MONEY = "#,##0.0;(#,##0.0);-"
PCT = "0.0%"


def _row(ws, r, label, values, font=FORMULA, fmt=MONEY, start_col=2):
    ws.cell(row=r, column=1, value=label).font = HEADER
    for i, v in enumerate(values):
        cell = ws.cell(row=r, column=start_col + i, value=v)
        cell.font = font
        cell.number_format = fmt


def build_cover(ws, aset: AssumptionSet, book: BookData) -> None:
    ws.cell(row=2, column=2, value=aset.profile.name).font = TITLE
    rows = [
        ("Тип проекта", aset.profile.project_type),
        ("Отрасль", aset.profile.industry),
        ("Локация", aset.profile.location),
        ("Валюта", aset.profile.currency),
        ("Горизонт, лет", aset.profile.horizon_years),
        ("Старт модели", aset.profile.model_start.isoformat() if aset.profile.model_start else None),
    ]
    r = 4
    for label, value in rows:
        if value is None:
            continue
        ws.cell(row=r, column=2, value=label).font = HEADER
        ws.cell(row=r, column=3, value=value).font = INPUT
        r += 1
    r += 1
    for q in aset.open_questions:
        ws.cell(row=r, column=2, value=("⛔ " if q.severity == "blocker" else "⚠ ") + q.question).font = NOTE
        r += 1
    ws.cell(row=r + 1, column=2,
            value="Книга собрана по схеме finmodel.assumptions.v1 (AFM&C)").font = NOTE
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 46


def build_assumptions(ws, aset: AssumptionSet, book: BookData) -> None:
    ws.cell(row=1, column=1, value="Допущения (вводы — синим; жёлтая заливка — ключевые)").font = HEADER
    r = 3
    ws.cell(row=r, column=1, value="Продукты").font = HEADER
    r += 1
    for p in aset.products:
        ws.cell(row=r, column=1, value=p.name).font = HEADER
        c_price = ws.cell(row=r, column=2, value=p.start_price)
        c_price.font = INPUT
        c_price.fill = KEY_FILL
        c_price.number_format = MONEY
        c_vol = ws.cell(row=r, column=3, value=p.start_volume)
        c_vol.font = INPUT
        c_vol.number_format = "#,##0"
        ws.cell(row=r, column=4, value=f"разгон {p.ramp_up_months or 0} мес").font = NOTE
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Финансирование").font = HEADER
    r += 1
    ws.cell(row=r, column=1, value="Собственные средства")
    equity = ws.cell(row=r, column=2, value=aset.financing.equity_amount)
    equity.font = INPUT
    equity.number_format = MONEY
    r += 1
    for f in aset.financing.facilities:
        ws.cell(row=r, column=1, value=f"Кредит: {f.name}").font = HEADER
        amount = ws.cell(row=r, column=2, value=f.amount)
        amount.font = INPUT
        amount.fill = KEY_FILL
        amount.number_format = MONEY
        ws.cell(row=r, column=3, value=f"{f.term_months} мес, грейс {f.grace_months}").font = NOTE
        rate = ws.cell(row=r, column=4, value=f.rate.points[0].value_pct / 100)
        rate.font = INPUT
        rate.number_format = PCT
        if f.rate_formula:
            ws.cell(row=r, column=5, value=f.rate_formula).font = NOTE
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Оценка").font = HEADER
    r += 1
    ws.cell(row=r, column=1, value="Ставка дисконтирования")
    disc = ws.cell(row=r, column=2, value=aset.valuation.discount_rate_pct / 100)
    disc.font = INPUT
    disc.fill = KEY_FILL
    disc.number_format = PCT
    if aset.valuation.reinvest_roa_pct is not None:
        r += 1
        ws.cell(row=r, column=1, value="ROA реинвестирования (MIRR)")
        roa = ws.cell(row=r, column=2, value=aset.valuation.reinvest_roa_pct / 100)
        roa.font = INPUT
        roa.number_format = PCT
    r += 2
    for path, src in aset.sources.items():
        ws.cell(row=r, column=1,
                value=f"{path}: {src.method.value}"
                      + (f", {src.document}" if src.document else "")
                      + (f" ({src.locator})" if src.locator else "")
                      + f", conf={src.confidence:.2f}").font = NOTE
        r += 1
    ws.column_dimensions["A"].width = 30


def build_cf(ws, aset: AssumptionSet, book: BookData) -> None:
    n = len(book.months)
    ws.cell(row=1, column=1, value="Денежный поток, помесячно").font = HEADER
    ws.cell(row=1, column=2 + n,
            value="Ряды 3–8 — расчёт движка AFM&C по Допущениям; итоги ниже — формулы").font = NOTE
    _row(ws, 2, "Месяц", [m.strftime("%Y-%m") for m in book.months], font=HEADER, fmt="@")
    _row(ws, 3, "Выручка", book.revenue)
    _row(ws, 4, "OPEX", book.opex)
    _row(ws, 5, "CAPEX", book.capex)
    _row(ws, 6, "Транш кредита", book.debt_draw)
    _row(ws, 7, "Проценты", book.interest)
    _row(ws, 8, "Погашение тела", book.principal)
    _row(ws, 9, "Налог на прибыль", book.tax)
    for i in range(n):
        col = get_column_letter(2 + i)
        ebitda = ws.cell(row=10, column=2 + i, value=f"={col}3-{col}4")
        ebitda.font = FORMULA
        ebitda.number_format = MONEY
        net = ws.cell(row=11, column=2 + i, value=f"={col}10-{col}9-{col}5+{col}6-{col}7-{col}8")
        net.font = FORMULA
        net.number_format = MONEY
        prev = get_column_letter(1 + i)
        cumulative = ws.cell(
            row=12, column=2 + i,
            value=f"={col}11" if i == 0 else f"={prev}12+{col}11")
        cumulative.font = FORMULA
        cumulative.number_format = MONEY
    ws.cell(row=10, column=1, value="EBITDA (формула)").font = HEADER
    ws.cell(row=11, column=1, value="Чистый поток после налога (формула)").font = HEADER
    ws.cell(row=12, column=1, value="Кумулятив (формула)").font = HEADER
    ws.column_dimensions["A"].width = 24
    ws.freeze_panes = "B3"


def build_dashboard(ws, aset: AssumptionSet, book: BookData) -> None:
    n = len(book.months)
    last = get_column_letter(1 + n)
    ws.cell(row=1, column=1, value="Дашборд").font = TITLE
    r = 3
    agg = [
        ("Выручка, итого", f"=SUM(CF!B3:{last}3)"),
        ("OPEX, итого", f"=SUM(CF!B4:{last}4)"),
        ("CAPEX, итого", f"=SUM(CF!B5:{last}5)"),
        ("Налог на прибыль, итого", f"=SUM(CF!B9:{last}9)"),
        ("EBITDA, итого", f"=SUM(CF!B10:{last}10)"),
        ("Чистый поток, итого", f"=SUM(CF!B11:{last}11)"),
    ]
    for label, formula in agg:
        ws.cell(row=r, column=1, value=label).font = HEADER
        cell = ws.cell(row=r, column=2, value=formula)
        cell.font = LINK
        cell.number_format = MONEY
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Инвестиционные метрики (движок AFM&C: FCFF, MIRR по ROA)").font = NOTE
    r += 1
    metric_labels = [
        ("NPV", book.metrics["npv"], MONEY),
        ("IRR", (book.metrics["irr_pct"] or 0) / 100, PCT),
        ("MIRR", (book.metrics["mirr_pct"] / 100) if book.metrics["mirr_pct"] is not None else None, PCT),
        ("Окупаемость, мес", book.metrics["payback_months"], "#,##0"),
        ("DSCR min", book.metrics["dscr_min"], "0.00"),
    ]
    for label, value, fmt in metric_labels:
        ws.cell(row=r, column=1, value=label).font = HEADER
        cell = ws.cell(row=r, column=2, value=value)
        cell.font = FORMULA
        cell.number_format = fmt
        r += 1
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18


def build_credit(ws, aset: AssumptionSet, book: BookData) -> None:
    n = len(book.months)
    _row(ws, 2, "Месяц", [m.strftime("%Y-%m") for m in book.months], font=HEADER, fmt="@")
    _row(ws, 3, "Проценты", book.interest)
    _row(ws, 4, "Погашение тела", book.principal)
    for i in range(n):
        col = get_column_letter(2 + i)
        cell = ws.cell(row=5, column=2 + i, value=f"={col}3+{col}4")
        cell.font = FORMULA
        cell.number_format = MONEY
    ws.cell(row=5, column=1, value="Обслуживание долга (формула)").font = HEADER
    _row(ws, 6, "Остаток долга", book.debt_outstanding)
    total = ws.cell(row=8, column=2, value=f"=SUM(B4:{get_column_letter(1 + n)}4)")
    total.font = FORMULA
    total.number_format = MONEY
    ws.cell(row=8, column=1, value="Погашено всего (формула)").font = HEADER
    ws.column_dimensions["A"].width = 26
    ws.freeze_panes = "B3"


def _series_header(ws, book: BookData, title: str) -> None:
    ws.cell(row=1, column=1, value=title).font = HEADER
    _row(ws, 2, "Месяц", [m.strftime("%Y-%m") for m in book.months], font=HEADER, fmt="@")
    ws.column_dimensions["A"].width = 28
    ws.freeze_panes = "B3"


def _formula_total(ws, r: int, n: int, label: str, first_row: int = 3) -> None:
    """Итоговая строка живой формулой SUM по колонке (не значением)."""
    for i in range(n):
        col = get_column_letter(2 + i)
        cell = ws.cell(row=r, column=2 + i,
                       value=f"=SUM({col}{first_row}:{col}{r - 1})" if r > first_row else 0)
        cell.font = FORMULA
        cell.number_format = MONEY
    ws.cell(row=r, column=1, value=label).font = HEADER


def build_sales(ws, aset: AssumptionSet, book: BookData) -> None:
    n = len(book.months)
    _series_header(ws, book, "План продаж — по продуктам (движок AFM&C)")
    r = 3
    for name, series in book.revenue_by_product.items():
        _row(ws, r, name, series)
        r += 1
    _formula_total(ws, max(r, 4), n, "Выручка итого (формула)")


def build_capex_sheet(ws, aset: AssumptionSet, book: BookData) -> None:
    _series_header(ws, book, "CAPEX и амортизация")
    _row(ws, 3, "CAPEX (график)", book.capex)
    _row(ws, 4, "Амортизация", book.depreciation)
    r = 6
    ws.cell(row=r, column=1, value="Позиции (вводы)").font = HEADER
    r += 1
    for item in aset.capex.items:
        ws.cell(row=r, column=1, value=item.name).font = HEADER
        cell = ws.cell(row=r, column=2, value=item.amount)
        cell.font = INPUT
        cell.fill = KEY_FILL
        cell.number_format = MONEY
        sm = item.schedule_months
        ws.cell(row=r, column=3, value=f"мес {sm[0]}–{sm[1]}" if sm else "мес 0").font = NOTE
        if item.vat_included:
            ws.cell(row=r, column=4, value="НДС внутри").font = NOTE
        r += 1
    if aset.capex.total_override is not None:
        ws.cell(row=r, column=1, value="Итог (без разбивки)").font = HEADER
        cell = ws.cell(row=r, column=2, value=aset.capex.total_override)
        cell.font = INPUT
        cell.number_format = MONEY
        r += 1
    if aset.capex.depreciation_months:
        ws.cell(row=r, column=1, value="Срок амортизации, мес")
        cell = ws.cell(row=r, column=2, value=aset.capex.depreciation_months)
        cell.font = INPUT
        cell.number_format = "#,##0"


def build_opex_sheet(ws, aset: AssumptionSet, book: BookData) -> None:
    n = len(book.months)
    _series_header(ws, book, "Операционные расходы — по статьям (фикс + % выручки)")
    r = 3
    for name, series in book.opex_by_item.items():
        _row(ws, r, name, series)
        r += 1
    _formula_total(ws, max(r, 4), n, "OPEX итого (формула)")


def build_pl(ws, aset: AssumptionSet, book: BookData) -> None:
    n = len(book.months)
    _series_header(ws, book, "Прибыли и убытки, помесячно")
    _row(ws, 3, "Выручка", book.revenue)
    _row(ws, 4, "OPEX", book.opex)
    for i in range(n):
        col = get_column_letter(2 + i)
        cell = ws.cell(row=5, column=2 + i, value=f"={col}3-{col}4")
        cell.font = FORMULA
        cell.number_format = MONEY
    ws.cell(row=5, column=1, value="EBITDA (формула)").font = HEADER
    _row(ws, 6, "Амортизация", book.depreciation)
    for i in range(n):
        col = get_column_letter(2 + i)
        cell = ws.cell(row=7, column=2 + i, value=f"={col}5-{col}6")
        cell.font = FORMULA
        cell.number_format = MONEY
    ws.cell(row=7, column=1, value="EBIT (формула)").font = HEADER
    _row(ws, 8, "Проценты по кредитам", book.interest)
    _row(ws, 9, "Налог на прибыль", book.tax)
    for i in range(n):
        col = get_column_letter(2 + i)
        cell = ws.cell(row=10, column=2 + i, value=f"={col}7-{col}8-{col}9")
        cell.font = FORMULA
        cell.number_format = MONEY
    ws.cell(row=10, column=1, value="Чистая прибыль (формула)").font = HEADER


def build_roadmap(ws, aset: AssumptionSet, book: BookData) -> None:
    ws.cell(row=1, column=1, value="Дорожная карта проекта").font = TITLE
    kind_labels = {"stage": "Этап", "capex": "CAPEX", "launch": "Запуск", "finance": "Финансирование"}
    r = 3
    ws.cell(row=r, column=1, value="Месяц").font = HEADER
    ws.cell(row=r, column=2, value="Дата").font = HEADER
    ws.cell(row=r, column=3, value="Тип").font = HEADER
    ws.cell(row=r, column=4, value="Веха").font = HEADER
    r += 1
    for m in sorted(aset.milestones, key=lambda x: x.month):
        ws.cell(row=r, column=1, value=m.month).font = INPUT
        if m.month < len(book.months):
            ws.cell(row=r, column=2, value=book.months[m.month].strftime("%Y-%m")).font = FORMULA
        ws.cell(row=r, column=3, value=kind_labels.get(m.kind, m.kind)).font = FORMULA
        ws.cell(row=r, column=4, value=m.name + (f" — {m.note}" if m.note else "")).font = INPUT
        r += 1
    # производные вехи из расчёта — связь карты с CAPEX и запуском продаж
    r += 1
    ws.cell(row=r, column=1, value="Расчётные вехи (движок)").font = HEADER
    r += 1
    derived = []
    capex_months = [i for i, v in enumerate(book.capex) if v > 0]
    if capex_months:
        derived.append((capex_months[0], "Начало инвестиционной фазы (CAPEX)"))
        derived.append((capex_months[-1], "Завершение CAPEX"))
    revenue_months = [i for i, v in enumerate(book.revenue) if v > 0]
    if revenue_months:
        derived.append((revenue_months[0], "Старт продаж"))
    positive = [i for i, v in enumerate(book.cumulative_cf) if v >= 0]
    if positive and any(v < 0 for v in book.cumulative_cf[:positive[-1] + 1]):
        derived.append((positive[0], "Выход накопленного потока в плюс"))
    for month, label in sorted(derived):
        ws.cell(row=r, column=1, value=month).font = FORMULA
        ws.cell(row=r, column=2, value=book.months[month].strftime("%Y-%m")).font = FORMULA
        ws.cell(row=r, column=4, value=label).font = NOTE
        r += 1
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 52


def build_staff(ws, aset: AssumptionSet, book: BookData) -> None:
    n = len(book.months)
    _series_header(ws, book, "ФОТ: план набора и расходы на персонал")
    r = 3
    for name, series in book.payroll_by_role.items():
        _row(ws, r, name, series)
        r += 1
    _formula_total(ws, max(r, 4), n, "ФОТ итого (формула)")
    r = max(r, 4) + 1
    contributions = book.opex_by_item.get("Страховые взносы")
    if contributions:
        _row(ws, r, "Страховые взносы", contributions)
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Штатное расписание (вводы)").font = HEADER
    r += 1
    for role in aset.staff.roles:
        ws.cell(row=r, column=1, value=role.name).font = HEADER
        count = ws.cell(row=r, column=2, value=role.count)
        count.font = INPUT
        count.number_format = "#,##0.#"
        salary = ws.cell(row=r, column=3, value=role.monthly_salary)
        salary.font = INPUT
        salary.number_format = MONEY
        span = f"с мес {role.start_month}" + (f" по {role.end_month}" if role.end_month is not None else "")
        ws.cell(row=r, column=4, value=span).font = NOTE
        r += 1


def build_balance(ws, aset: AssumptionSet, book: BookData) -> None:
    n = len(book.months)
    _series_header(ws, book, "Прогнозный баланс (упрощённый), на конец месяца")
    _row(ws, 3, "Внеоборотные активы (ОС)", book.fixed_assets)
    _row(ws, 4, "Денежные средства", book.cash)
    for i in range(n):
        col = get_column_letter(2 + i)
        cell = ws.cell(row=5, column=2 + i, value=f"={col}3+{col}4")
        cell.font = FORMULA
        cell.number_format = MONEY
    ws.cell(row=5, column=1, value="АКТИВЫ (формула)").font = HEADER
    _row(ws, 6, "Собственный капитал", book.equity_book)
    _row(ws, 7, "Долг", book.debt_outstanding)
    for i in range(n):
        col = get_column_letter(2 + i)
        cell = ws.cell(row=8, column=2 + i, value=f"={col}6+{col}7")
        cell.font = FORMULA
        cell.number_format = MONEY
        check = ws.cell(row=9, column=2 + i, value=f"={col}5-{col}8")
        check.font = FORMULA
        check.number_format = MONEY
    ws.cell(row=8, column=1, value="ПАССИВЫ (формула)").font = HEADER
    ws.cell(row=9, column=1, value="Проверка: Активы − Пассивы = 0").font = NOTE
    ws.cell(row=11, column=1,
            value="Деньги включают взнос собственного капитала на старте; "
                  "капитал прирастает чистой прибылью. Оборотный капитал — в развитии.").font = NOTE


def build_covenants(ws, aset: AssumptionSet, book: BookData) -> None:
    ws.cell(row=1, column=1, value="Ковенанты и долговые метрики").font = TITLE
    cov = aset.financing.covenants
    metrics = book.metrics
    years = len(book.months) // 12
    r = 3
    ws.cell(row=r, column=1, value="Метрика").font = HEADER
    for y in range(years):
        ws.cell(row=r, column=2 + y, value=f"Год {y + 1}").font = HEADER
    ws.cell(row=r, column=2 + years, value="Порог").font = HEADER
    ws.cell(row=r, column=3 + years, value="Статус").font = HEADER
    rows = [
        ("DSCR", metrics.get("dscr_by_year", []), cov.dscr_min, "≥"),
        ("ICR (EBITDA/проценты)", metrics.get("icr_by_year", []), cov.icr_min, "≥"),
        ("Долг/EBITDA", metrics.get("net_debt_to_ebitda_by_year", []), cov.net_debt_to_ebitda_max, "≤"),
    ]
    r += 1
    for label, values, threshold, op in rows:
        ws.cell(row=r, column=1, value=label).font = HEADER
        for y, value in enumerate(values):
            cell = ws.cell(row=r, column=2 + y, value=value)
            cell.font = FORMULA
            cell.number_format = "0.00"
        if threshold is not None:
            th = ws.cell(row=r, column=2 + years, value=f"{op} {threshold}")
            th.font = INPUT
            th.fill = KEY_FILL
            if values:
                worst = min(values) if op == "≥" else max(values)
                ok = worst >= threshold if op == "≥" else worst <= threshold
                ws.cell(row=r, column=3 + years, value="OK" if ok else "НАРУШЕНИЕ").font = (
                    FORMULA if ok else Font(name=FONT, bold=True, color="CC0000"))
        r += 1
    r += 1
    breaches = metrics.get("covenant_breaches", [])
    if breaches:
        ws.cell(row=r, column=1, value="Нарушения:").font = HEADER
        r += 1
        for b in breaches:
            ws.cell(row=r, column=1, value="⛔ " + b).font = NOTE
            r += 1
    elif any(t is not None for _, _, t, _ in rows):
        ws.cell(row=r, column=1, value="Все заданные ковенанты выполняются").font = NOTE
    else:
        ws.cell(row=r, column=1, value="Пороги ковенант не заданы (financing.covenants)").font = NOTE
    ws.column_dimensions["A"].width = 26


SENSITIVITY_FACTORS = [0.8, 0.9, 1.0, 1.1, 1.2]


def build_sensitivity(ws, aset: AssumptionSet, book: BookData) -> None:
    ws.cell(row=1, column=1, value="Чувствительность NPV: цена × объём продаж").font = TITLE
    factors = SENSITIVITY_FACTORS
    matrix = npv_sensitivity(aset, factors, factors)
    ws.cell(row=3, column=1, value="Цена \\ Объём").font = HEADER
    for j, vf in enumerate(factors):
        cell = ws.cell(row=3, column=2 + j, value=vf - 1)
        cell.font = HEADER
        cell.number_format = "+0%;-0%;0%"
    for i, pf in enumerate(factors):
        cell = ws.cell(row=4 + i, column=1, value=pf - 1)
        cell.font = HEADER
        cell.number_format = "+0%;-0%;0%"
        for j in range(len(factors)):
            value = ws.cell(row=4 + i, column=2 + j, value=matrix[i][j])
            value.font = FORMULA
            value.number_format = MONEY
    last_col = get_column_letter(1 + len(factors))
    ws.conditional_formatting.add(
        f"B4:{last_col}{3 + len(factors)}",
        ColorScaleRule(start_type="min", start_color="F8696B",
                       mid_type="num", mid_value=0, mid_color="FFEB84",
                       end_type="max", end_color="63BE7B"))
    ws.cell(row=5 + len(factors), column=1,
            value="Центр матрицы — базовый сценарий; расчёт движка AFM&C по всем продуктам").font = NOTE
    ws.column_dimensions["A"].width = 14


# Контракт именованных ячеек для внешних агентов (ИИ-агент работает в Excel
# по стабильным именам, а не по адресам). Часть стандарта AFM&C.
_CF_NAMED_ROWS = [
    ("RevenueRow", 3), ("OpexRow", 4), ("CapexRow", 5), ("DebtDrawRow", 6),
    ("InterestRow", 7), ("PrincipalRow", 8), ("ProfitTaxRow", 9),
    ("EbitdaRow", 10), ("NetCFRow", 11), ("CumCFRow", 12),
]
_DASHBOARD_NAMED_METRICS = [
    ("NPV", 11), ("IRR", 12), ("MIRR", 13), ("PaybackMonths", 14), ("DSCRmin", 15),
]


def _add_defined_names(workbook: Workbook, specs: list[SheetSpec], n_months: int) -> None:
    titles = {s.code: s.title for s in specs}
    if "cf" in titles:
        last = get_column_letter(1 + n_months)
        for name, row in _CF_NAMED_ROWS:
            workbook.defined_names.add(DefinedName(
                name, attr_text=f"'{titles['cf']}'!$B${row}:${last}${row}"))
    if "dashboard" in titles:
        for name, row in _DASHBOARD_NAMED_METRICS:
            workbook.defined_names.add(DefinedName(
                name, attr_text=f"'{titles['dashboard']}'!$B${row}"))


_BUILDERS = {
    "build_cover": build_cover,
    "build_roadmap": build_roadmap,
    "build_staff": build_staff,
    "build_balance": build_balance,
    "build_covenants": build_covenants,
    "build_assumptions": build_assumptions,
    "build_cf": build_cf,
    "build_dashboard": build_dashboard,
    "build_sales": build_sales,
    "build_capex_sheet": build_capex_sheet,
    "build_opex_sheet": build_opex_sheet,
    "build_pl": build_pl,
    "build_credit": build_credit,
    "build_sensitivity": build_sensitivity,
}


def export_book_xlsx(aset: AssumptionSet, book: BookData,
                     sheet_codes: list[str] | None = None) -> bytes:
    """Собирает книгу по реестру и возвращает xlsx-байты."""
    specs: list[SheetSpec] = resolve_sheets(sheet_codes)
    workbook = Workbook()
    workbook.remove(workbook.active)
    for spec in specs:
        ws = workbook.create_sheet(title=spec.title)
        _BUILDERS[spec.builder_name](ws, aset, book)
        for row in ws.iter_rows():
            for cell in row:
                if cell.font is None or cell.font.name != FONT:
                    cell.font = Font(name=FONT,
                                     bold=cell.font.bold if cell.font else False,
                                     italic=cell.font.italic if cell.font else False,
                                     size=cell.font.size if cell.font else 11,
                                     color=cell.font.color if cell.font else None)
        ws.sheet_view.showGridLines = False
    _add_defined_names(workbook, specs, len(book.months))
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
