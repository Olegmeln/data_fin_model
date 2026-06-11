"""Парсер банковских выписок в формате Excel (.xlsx)."""
import io

from openpyxl import load_workbook

from .base import ParsedOperation, ParserError, build_operation, map_columns


def parse_xlsx(raw: bytes) -> list[ParsedOperation]:
    try:
        workbook = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ParserError(f"Не удалось открыть файл Excel: {exc}") from exc

    sheet = workbook.active
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    workbook.close()
    rows = [row for row in rows if any(c is not None and str(c).strip() for c in row)]
    if not rows:
        raise ParserError("Лист Excel пуст.")

    header_index = None
    for i, row in enumerate(rows[:10]):
        joined = " ".join(str(c).lower() for c in row if c is not None)
        if "дата" in joined or "date" in joined:
            header_index = i
            break
    if header_index is None:
        raise ParserError("Не найдена строка заголовков: ожидается колонка «Дата».")

    cols = map_columns(rows[header_index])
    if "date" not in cols or ("amount" not in cols and "debit" not in cols and "credit" not in cols):
        raise ParserError("Не удалось определить колонки даты и суммы.")

    operations: list[ParsedOperation] = []
    for row in rows[header_index + 1:]:
        op = build_operation(row, cols)
        if op is not None:
            operations.append(op)

    if not operations:
        raise ParserError("В файле не найдено ни одной операции.")
    return operations
