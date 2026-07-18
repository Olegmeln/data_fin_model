"""Тесты реестра листов, генерации книги и API экспорта."""
import io
from datetime import date

import pytest
from openpyxl import load_workbook

import app.api as api_module
from app.finmodel.assumptions_schema import (
    AssumptionSet, Capex, CapexItem, CreditFacility, Opex, OpexItem, Product,
    ProjectProfile, RateSchedule, Scenario, Valuation,
)
from app.finmodel.book import build_book
from app.finmodel.book.excel import export_book_xlsx
from app.finmodel.book.registry import REGISTRY, resolve_sheets


def book_set() -> AssumptionSet:
    return AssumptionSet(
        profile=ProjectProfile(name="Книга-тест", horizon_years=2, model_start=date(2026, 1, 1)),
        products=[Product(name="Товар", start_price=10, start_volume=100, ramp_up_months=2)],
        opex=Opex(items=[OpexItem(name="Аренда", monthly_amount=100)]),
        capex=Capex(items=[CapexItem(name="Станок", amount=600)]),
        financing={"equity_amount": 200, "facilities": [
            CreditFacility(name="Кредит", amount=400, term_months=12, grace_months=3,
                           rate=RateSchedule.flat(12))]},
        valuation=Valuation(discount_rate_pct=10, reinvest_roa_pct=15),
        scenarios=[Scenario(name="Стресс", overrides={"products.0.start_volume": 50})],
    )


class TestRegistry:
    def test_default_composition_and_order(self):
        codes = [s.code for s in resolve_sheets()]
        assert codes == ["cover", "assumptions", "cf", "dashboard", "credit"]

    def test_subset_keeps_requested_order(self):
        codes = [s.code for s in resolve_sheets(["dashboard", "cf"])]
        assert codes == ["dashboard", "cf"]

    def test_unknown_code_raises_with_known_list(self):
        with pytest.raises(ValueError, match="неизвестные листы"):
            resolve_sheets(["cf", "wat"])

    def test_codes_unique(self):
        codes = [s.code for s in REGISTRY]
        assert len(codes) == len(set(codes))


class TestExcelBook:
    def _workbook(self, sheets=None):
        aset = book_set()
        payload = export_book_xlsx(aset, build_book(aset), sheets)
        return load_workbook(io.BytesIO(payload))

    def test_all_sheets_present(self):
        assert self._workbook().sheetnames == ["Обложка", "Допущения", "CF", "Дашборд", "Кредит"]

    def test_subset_export(self):
        assert self._workbook(["cf", "dashboard"]).sheetnames == ["CF", "Дашборд"]

    def test_cf_totals_are_formulas_not_values(self):
        cf = self._workbook()["CF"]
        assert str(cf.cell(row=9, column=2).value).startswith("=")   # EBITDA
        assert str(cf.cell(row=10, column=2).value).startswith("=")  # чистый поток
        assert str(cf.cell(row=11, column=3).value).startswith("=B11")  # кумулятив тянет предыдущий

    def test_dashboard_links_cf_by_formula(self):
        dash = self._workbook()["Дашборд"]
        assert "CF!" in str(dash.cell(row=3, column=2).value)

    def test_month_headers_span_horizon(self):
        cf = self._workbook()["CF"]
        assert cf.cell(row=2, column=2).value == "2026-01"
        assert cf.cell(row=2, column=25).value == "2027-12"

    def test_cover_carries_open_questions(self):
        aset = book_set()
        aset.open_questions.append(
            __import__("app.finmodel.assumptions_schema", fromlist=["OpenQuestion"]).OpenQuestion(
                question="Проверочный вопрос", severity="blocker"))
        payload = export_book_xlsx(aset, build_book(aset))
        cover = load_workbook(io.BytesIO(payload))["Обложка"]
        texts = [c.value for row in cover.iter_rows() for c in row if isinstance(c.value, str)]
        assert any("Проверочный вопрос" in t for t in texts)


class TestBookApi:
    def _seed(self, client, project="bookpr"):
        body = {"assumptions": book_set().model_dump(by_alias=True, exclude_none=True, mode="json"),
                "status": "draft"}
        response = client.put(f"/api/assumption-sets/{project}", json=body)
        assert response.status_code == 200, response.text

    def test_metrics_endpoint(self, client):
        self._seed(client)
        payload = client.get("/api/book/bookpr/metrics").json()
        assert payload["metrics"]["npv"] is not None
        assert payload["metrics"]["mirr_pct"] is not None

    def test_metrics_scenario_changes_result(self, client):
        self._seed(client, "bookpr2")
        base = client.get("/api/book/bookpr2/metrics").json()["metrics"]["revenue_total"]
        stress = client.get("/api/book/bookpr2/metrics",
                            params={"scenario": "Стресс"}).json()["metrics"]["revenue_total"]
        assert stress == pytest.approx(base * 0.5)

    def test_metrics_unknown_scenario_422(self, client):
        self._seed(client, "bookpr3")
        assert client.get("/api/book/bookpr3/metrics",
                          params={"scenario": "нет"}).status_code == 422

    def test_export_returns_valid_workbook(self, client):
        self._seed(client, "bookpr4")
        response = client.get("/api/book/bookpr4/export")
        assert response.status_code == 200
        workbook = load_workbook(io.BytesIO(response.content))
        assert "Дашборд" in workbook.sheetnames
        assert "attachment" in response.headers["content-disposition"]

    def test_export_sheet_filter(self, client):
        self._seed(client, "bookpr5")
        response = client.get("/api/book/bookpr5/export", params={"sheets": "cf,dashboard"})
        assert load_workbook(io.BytesIO(response.content)).sheetnames == ["CF", "Дашборд"]

    def test_export_unknown_project_404(self, client):
        assert client.get("/api/book/absent/export").status_code == 404
